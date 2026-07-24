import io
import openpyxl
from apps.quizzes.models import QuestionBank, QuizQuestion, QuizQuestionOption


class ExcelQuizImporter:
    @staticmethod
    def generate_template_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"

        headers = [
            "Savol matni", "Turi (single/multi)", "Bali", "Manfiy bal",
            "1-Variant", "1-To'g'ri (1/0)", "2-Variant", "2-To'g'ri (1/0)",
            "3-Variant", "3-To'g'ri (1/0)", "4-Variant", "4-To'g'ri (1/0)"
        ]
        ws.append(headers)

        sample_row1 = [
            "O'zbekiston poytaxti qaysi shahar?", "single", 1.0, 0.0,
            "Toshkent", 1, "Samarqand", 0, "Buxoro", 0, "Xiva", 0
        ]
        sample_row2 = [
            "Qaysi sonlar juft?", "multi", 2.0, 0.0,
            "2", 1, "3", 0, "4", 1, "5", 0
        ]
        ws.append(sample_row1)
        ws.append(sample_row2)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def import_from_excel(file_obj, question_bank: QuestionBank):
        wb = openpyxl.load_workbook(file_obj)
        sheet = wb.active

        created_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            q_text = str(row[0]).strip()

            def parse_float(val, default=1.0):
                if val is None:
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default

            col_1_str = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ''

            if col_1_str in ['single', 'multi', 'text']:
                q_type = col_1_str
                points = parse_float(row[2] if len(row) > 2 else None, default=1.0)
                neg_marking = parse_float(row[3] if len(row) > 3 else None, default=0.0)
                start_opt_idx = 4
                step = 2
            else:
                q_type = 'single'
                points = parse_float(row[2] if len(row) > 2 else None, default=1.0)
                neg_marking = parse_float(row[3] if len(row) > 3 else None, default=0.0)
                start_opt_idx = 4 if len(row) > 4 else 1
                step = 2 if len(row) > 4 else 1

            question = QuizQuestion.objects.create(
                bank=question_bank,
                text=q_text,
                type=q_type,
                points=points,
                negative_marking=neg_marking
            )

            col_idx = start_opt_idx
            while col_idx < len(row):
                opt_text = row[col_idx]
                if opt_text is not None and str(opt_text).strip():
                    is_correct = False
                    if step == 2 and col_idx + 1 < len(row):
                        is_correct_val = row[col_idx + 1]
                        if is_correct_val in [True, 1, '1', 'true', 'True', 'YES', 'yes', 'to\'g\'ri', 'togri']:
                            is_correct = True

                    QuizQuestionOption.objects.create(
                        question=question,
                        text=str(opt_text).strip(),
                        is_correct=is_correct
                    )
                col_idx += step

            created_count += 1

        return created_count
